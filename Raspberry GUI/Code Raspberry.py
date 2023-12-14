import customtkinter as ctk
from PIL import Image, ImageTk
import time
import datetime
import threading
from openpyxl import load_workbook




# Déclaration des variables globales initiales et de sauvegarde
slider_value = 8
slider_value2 = 16
saved_slider_value = 8
saved_slider_value2 = 16
selected_on_hour = 0
selected_off_hour = 0
saved_on_hour = 0
saved_off_hour = 0
slider_value3= 0
saved_intensity=0

# Définition des variables globales pour les valeurs des sliders et des comboboxes pour la LED n°2
slider_value_led2 = 8
slider_value2_led2 = 16
saved_slider_value_led2 = 8
saved_slider_value2_led2 = 16
selected_on_hour_led2 = 0
selected_off_hour_led2 = 0
saved_on_hour_led2 = 0
saved_off_hour_led2 = 0
slider_value3_led2= 0
saved_intensity_led2=0

# Définition des variables globales pour les valeurs des sliders et des comboboxes pour la LED n°3
slider_value_led3 = 8
slider_value2_led3 = 16
saved_slider_value_led3 = 8
saved_slider_value2_led3 = 16
selected_on_hour_led3 = 0
selected_off_hour_led3 = 0
saved_on_hour_led3 = 0
saved_off_hour_led3 = 0
slider_value3_led3= 0
saved_intensity_led3=0

# Déclaration des variables globales pour stocker les valeurs sauvegardées
saved_minutes_on = "00"
saved_seconds_on = "00"
saved_minutes_off = "00"
saved_seconds_off = "00"

# Déclaration des variables globales pour les sliders
slider_pump2 = None
slider2_pump2 = None

# Variables globales pour stocker les valeurs sauvegardées
saved_slider_value_pump2_on = 0
saved_slider_value_pump2_off = 0

def update_time():
    current_time = time.strftime('%H:%M:%S')  # Format de l'heure : heures:minutes:secondes
    time_label.configure(text=current_time)
    home_frame.after(1000, update_time)  # Mise à jour de l'heure toutes les secondes

def send_system1(heure_on, heure_off,intensity):
    # Chemin du fichier Excel existant
    fichier_excel = 'Aquapo.xlsx'

    # Charger le workbook (classeur)
    workbook = load_workbook(fichier_excel)

    # Sélectionner la feuille de calcul active ou par son nom
    feuille = workbook.active
    # Ou si vous connaissez le nom de la feuille
    # feuille = workbook['NomDeLaFeuille']
    feuille['B2'] = heure_on
    feuille['B3'] = heure_off
    feuille['B4'] = intensity
    workbook.save(fichier_excel)

def send_system2(heure_on, heure_off,intensity):
    # Chemin du fichier Excel existant
    fichier_excel = 'Aquapo.xlsx'

    # Charger le workbook (classeur)
    workbook = load_workbook(fichier_excel)

    # Sélectionner la feuille de calcul active ou par son nom
    feuille = workbook.active
    # Ou si vous connaissez le nom de la feuille
    # feuille = workbook['NomDeLaFeuille']
    feuille['C2'] = heure_on
    feuille['C3'] = heure_off
    feuille['C4'] = intensity
    workbook.save(fichier_excel)

def send_system3(heure_on, heure_off):
    # Chemin du fichier Excel existant
    fichier_excel = 'Aquapo.xlsx'

    # Charger le workbook (classeur)
    workbook = load_workbook(fichier_excel)

    # Sélectionner la feuille de calcul active ou par son nom
    feuille = workbook.active
    # Ou si vous connaissez le nom de la feuille
    # feuille = workbook['NomDeLaFeuille']
    feuille['D2'] = heure_on
    feuille['D3'] = heure_off
    workbook.save(fichier_excel)

def send_system4(min_on,min_off):
    # Chemin du fichier Excel existant
    fichier_excel = 'Aquapo.xlsx'

    # Charger le workbook (classeur)
    workbook = load_workbook(fichier_excel)

    # Sélectionner la feuille de calcul active ou par son nom
    feuille = workbook.active
    # Ou si vous connaissez le nom de la feuille
    # feuille = workbook['NomDeLaFeuille']
    feuille['F2'] = min_on
    feuille['F3'] = min_off
    workbook.save(fichier_excel)

def save_values():
    global saved_slider_value, saved_slider_value2, saved_on_hour, saved_off_hour, saved_intensity
    saved_slider_value = slider_value
    saved_slider_value2 = slider_value2
    saved_on_hour = selected_on_hour
    saved_off_hour = selected_off_hour
    save_intensity= slider_value3
    send_system1(selected_on_hour,selected_off_hour,slider_value3)

def show_led1_management():
    global slider_value, slider_value2, selected_on_hour, selected_off_hour, slider_value3

    # Effacer tout contenu précédent dans home_frame
    for widget in home_frame.winfo_children():
        widget.destroy()

    # Ajouter le titre
    led1_title = ctk.CTkLabel(home_frame, text="LED n°1 Management", font=("Roboto", 36))
    led1_title.pack(pady=20)

    # Ajouter le texte "ON TIME (hours)"
    on_time_label = ctk.CTkLabel(home_frame, text="ON TIME (hours)", font=("Roboto", 18))
    on_time_label.pack(anchor='w', padx=180, pady=10)

    # Fonctions de mise à jour des sliders et des labels
    def update_label(value):
        global slider_value, slider_value2
        slider_value = int(value)
        slider_value2 = 24 - slider_value
        slider2.set(slider_value2)
        slider_value_label.configure(text=f"{slider_value} Hours")
        slider_value_label2.configure(text=f"{slider_value2} Hours")
    
    def update_label2(value):
        global slider_value, slider_value2
        slider_value2 = int(value)
        slider_value = 24 - slider_value2
        slider.set(slider_value)
        slider_value_label.configure(text=f"{slider_value} Hours")
        slider_value_label2.configure(text=f"{slider_value2} Hours")

    # Créer les sliders
    slider = ctk.CTkSlider(home_frame, from_=0, to=24, number_of_steps=24, command=update_label)
    slider.pack(pady=10)

    slider_value_label = ctk.CTkLabel(home_frame, text="0 Hours")
    slider_value_label.place(x=570, y=132)

     # Ajouter le texte "ON TIME (hours)"
    off_time_label = ctk.CTkLabel(home_frame, text="OFF TIME (hours)", font=("Roboto", 18))
    off_time_label.pack(anchor='w', padx=180, pady=10)

    slider2 = ctk.CTkSlider(home_frame, from_=0, to=24, number_of_steps=24, command=update_label2)
    slider2.pack(pady=10)

    slider_value_label2 = ctk.CTkLabel(home_frame, text="0 Hours")
    slider_value_label2.place(x=570, y=220)

    # Ajout de labels pour "Start On Time" et "Start Off Time" avec couleur de texte spécifiée
    start_on_time_label = ctk.CTkLabel(home_frame, text="Start On Time", font=("Roboto", 18))
    start_on_time_label.pack(pady=20)
    
    def on_time_combobox_changed(heure):
        global selected_on_hour, selected_off_hour
        selected_on_hour = int(heure[:-1])
        selected_off_hour = (selected_on_hour + slider_value) % 24
        off_time_combobox.set(f"{selected_off_hour:02d}h")
    
    def off_time_combobox_changed(heure):
        global selected_on_hour, selected_off_hour
        selected_off_hour = int(heure[:-1])
        selected_on_hour = (selected_off_hour + slider_value2) % 24
        on_time_combobox.set(f"{selected_on_hour:02d}h")

    heures = [f"{i:02d}h" for i in range(24)]

    # Création des comboboxes et liaison des fonctions de callback
    on_time_combobox = ctk.CTkComboBox(home_frame, values=heures,command=on_time_combobox_changed)
    on_time_combobox.pack(pady=10)

    start_off_time_label = ctk.CTkLabel(home_frame, text="Start Off Time", font=("Roboto", 18))
    start_off_time_label.pack(pady=20)

    off_time_combobox = ctk.CTkComboBox(home_frame, values=heures,command=off_time_combobox_changed)
    off_time_combobox.pack(pady=10)

    def update_label3(value):
        global slider_value3
        slider_value3 = int(value)
        slider3.set(slider_value3)
        slider_value_label3.configure(text=f"{slider_value3}")

    # Ajouter le texte "Intensity"
    intensity_label = ctk.CTkLabel(home_frame, text="Intensity", font=("Roboto", 18))
    intensity_label.pack(anchor='w', padx=180, pady=10)

    slider3 = ctk.CTkSlider(home_frame, from_=0, to=10, number_of_steps=10, command=update_label3)
    slider3.pack(pady=10)

    slider_value_label3 = ctk.CTkLabel(home_frame, text="0")
    slider_value_label3.place(x=570, y=540)

    slider.set(slider_value)
    slider2.set(slider_value2)
    slider3.set(slider_value3)
    on_time_combobox.set(f"{selected_on_hour:02d}h")
    off_time_combobox.set(f"{selected_off_hour:02d}h")
    slider_value_label.configure(text=f"{slider_value} Hours")
    slider_value_label2.configure(text=f"{slider_value2} Hours")
    slider_value_label3.configure(text=f"{slider_value3}")


    # Création du bouton Save
    save_button = ctk.CTkButton(home_frame, text="Save",font=("Roboto", 24), command=save_values)
    save_button.pack(anchor='e',padx=100)


     # Afficher l'heure actuelle
    global time_label
    time_label = ctk.CTkLabel(home_frame, font=("Roboto", 18))
    time_label.place(x=820)
    update_time()
    
def save_values_led2():
    global saved_slider_value_led2, saved_slider_value2_led2, saved_on_hour_led2, saved_off_hour_led2, saved_intensity_led2
    saved_slider_value_led2 = slider_value_led2
    saved_slider_value2_led2 = slider_value2_led2
    saved_on_hour_led2 = selected_on_hour_led2
    saved_off_hour_led2 = selected_off_hour_led2
    save_intensity_led2= slider_value3_led2
    send_system2(selected_on_hour_led2,selected_off_hour_led2,slider_value3_led2)

def show_led2_management():
    global slider_value_led2, slider_value2_led2, selected_on_hour_led2, selected_off_hour_led2

    # Effacer tout contenu précédent dans home_frame
    for widget in home_frame.winfo_children():
        widget.destroy()

    # Ajouter le titre
    led2_title = ctk.CTkLabel(home_frame, text="LED n°2 Management", font=("Roboto", 36))
    led2_title.pack(pady=20)

    # Ajouter le texte "ON TIME (hours)"
    on_time_label = ctk.CTkLabel(home_frame, text="ON TIME (hours)", font=("Roboto", 18))
    on_time_label.pack(anchor='w', padx=180, pady=10)

    # Fonctions de mise à jour des sliders et des labels
    def update_label(value):
        global slider_value_led2, slider_value2_led2
        slider_value_led2 = int(value)
        slider_value2_led2 = 24 - slider_value_led2
        slider2.set(slider_value2_led2)
        slider_value_label.configure(text=f"{slider_value_led2} Hours")
        slider_value_label2.configure(text=f"{slider_value2_led2} Hours")
    
    def update_label2(value):
        global slider_value_led2, slider_value2_led2
        slider_value2_led2 = int(value)
        slider_value_led2 = 24 - slider_value2_led2
        slider.set(slider_value_led2)
        slider_value_label.configure(text=f"{slider_value_led2} Hours")
        slider_value_label2.configure(text=f"{slider_value2_led2} Hours")

    # Créer les sliders
    slider = ctk.CTkSlider(home_frame, from_=0, to=24, number_of_steps=24, command=update_label)
    slider.pack(pady=10)

    slider_value_label = ctk.CTkLabel(home_frame, text="0 Hours")
    slider_value_label.place(x=570, y=132)

     # Ajouter le texte "ON TIME (hours)"
    off_time_label = ctk.CTkLabel(home_frame, text="OFF TIME (hours)", font=("Roboto", 18))
    off_time_label.pack(anchor='w', padx=180, pady=10)

    slider2 = ctk.CTkSlider(home_frame, from_=0, to=24, number_of_steps=24, command=update_label2)
    slider2.pack(pady=10)

    slider_value_label2 = ctk.CTkLabel(home_frame, text="0 Hours")
    slider_value_label2.place(x=570, y=220)

    # Ajout de labels pour "Start On Time" et "Start Off Time" avec couleur de texte spécifiée
    start_on_time_label = ctk.CTkLabel(home_frame, text="Start On Time", font=("Roboto", 18))
    start_on_time_label.pack(pady=20)
    
    def on_time_combobox_changed(heure):
        global selected_on_hour_led2, selected_off_hour_led2
        selected_on_hour_led2 = int(heure[:-1])
        selected_off_hour_led2 = (selected_on_hour_led2 + slider_value_led2) % 24
        off_time_combobox.set(f"{selected_off_hour_led2:02d}h")
    
    def off_time_combobox_changed(heure):
        global selected_on_hour_led2, selected_off_hour_led2
        selected_off_hour_led2 = int(heure[:-1])
        selected_on_hour_led2 = (selected_off_hour_led2 + slider_value2_led2) % 24
        on_time_combobox.set(f"{selected_on_hour_led2:02d}h")

    heures = [f"{i:02d}h" for i in range(24)]

    # Création des comboboxes et liaison des fonctions de callback
    on_time_combobox = ctk.CTkComboBox(home_frame, values=heures,command=on_time_combobox_changed)
    on_time_combobox.pack(pady=10)

    start_off_time_label = ctk.CTkLabel(home_frame, text="Start Off Time", font=("Roboto", 18))
    start_off_time_label.pack(pady=20)

    off_time_combobox = ctk.CTkComboBox(home_frame, values=heures,command=off_time_combobox_changed)
    off_time_combobox.pack(pady=10)
    
    def update_label3(value):
        global slider_value3_led2
        slider_value3_led2 = int(value)
        slider3.set(slider_value3_led2)
        slider_value_label3.configure(text=f"{slider_value3_led2}")

    # Ajouter le texte "Intensity"
    intensity_label_led2 = ctk.CTkLabel(home_frame, text="Intensity", font=("Roboto", 18))
    intensity_label_led2.pack(anchor='w', padx=180, pady=10)

    slider3 = ctk.CTkSlider(home_frame, from_=0, to=10, number_of_steps=10, command=update_label3)
    slider3.pack(pady=10)

    slider_value_label3 = ctk.CTkLabel(home_frame, text="0")
    slider_value_label3.place(x=570, y=540)

    slider.set(slider_value_led2)
    slider2.set(slider_value2_led2)
    slider3.set(slider_value3_led2)
    on_time_combobox.set(f"{selected_on_hour_led2:02d}h")
    off_time_combobox.set(f"{selected_off_hour_led2:02d}h")
    slider_value_label.configure(text=f"{slider_value_led2} Hours")
    slider_value_label2.configure(text=f"{slider_value2_led2} Hours")

    # Création du bouton Save
    save_button = ctk.CTkButton(home_frame, text="Save",font=("Roboto", 24), command=save_values_led2)
    save_button.pack(anchor='e',padx=100)

     # Afficher l'heure actuelle
    global time_label
    time_label = ctk.CTkLabel(home_frame, font=("Roboto", 18))
    time_label.place(x=820)
    update_time()
    
def save_values_led3():
    global saved_slider_value_led3, saved_slider_value2_led3, saved_on_hour_led3, saved_off_hour_led3
    saved_slider_value_led3 = slider_value_led3
    saved_slider_value2_led3 = slider_value2_led3
    saved_on_hour_led3 = selected_on_hour_led3
    saved_off_hour_led3 = selected_off_hour_led3
    send_system3(selected_on_hour_led3,selected_off_hour_led3)

def show_led3_management():
    global slider_value_led3, slider_value2_led3, selected_on_hour_led3, selected_off_hour_led3

    # Effacer tout contenu précédent dans home_frame
    for widget in home_frame.winfo_children():
        widget.destroy()

    # Ajouter le titre
    led3_title = ctk.CTkLabel(home_frame, text="LED n°3 Management", font=("Roboto", 36))
    led3_title.pack(pady=20)

    # Ajouter le texte "ON TIME (hours)"
    on_time_label = ctk.CTkLabel(home_frame, text="ON TIME (hours)", font=("Roboto", 18))
    on_time_label.pack(anchor='w', padx=180, pady=10)

    # Fonctions de mise à jour des sliders et des labels
    def update_label(value):
        global slider_value_led3, slider_value2_led3
        slider_value_led3 = int(value)
        slider_value2_led3 = 24 - slider_value_led3
        slider2.set(slider_value2_led3)
        slider_value_label.configure(text=f"{slider_value_led3} Hours")
        slider_value_label2.configure(text=f"{slider_value2_led3} Hours")
    
    def update_label2(value):
        global slider_value_led3, slider_value2_led3
        slider_value2_led3 = int(value)
        slider_value_led3 = 24 - slider_value2_led3
        slider.set(slider_value_led3)
        slider_value_label.configure(text=f"{slider_value_led3} Hours")
        slider_value_label2.configure(text=f"{slider_value2_led3} Hours")

    # Créer les sliders
    slider = ctk.CTkSlider(home_frame, from_=0, to=24, number_of_steps=24, command=update_label)
    slider.pack(pady=10)

    slider_value_label = ctk.CTkLabel(home_frame, text="0 Hours")
    slider_value_label.place(x=570, y=132)

     # Ajouter le texte "ON TIME (hours)"
    off_time_label = ctk.CTkLabel(home_frame, text="OFF TIME (hours)", font=("Roboto", 18))
    off_time_label.pack(anchor='w', padx=180, pady=10)

    slider2 = ctk.CTkSlider(home_frame, from_=0, to=24, number_of_steps=24, command=update_label2)
    slider2.pack(pady=10)

    slider_value_label2 = ctk.CTkLabel(home_frame, text="0 Hours")
    slider_value_label2.place(x=570, y=220)

    # Ajout de labels pour "Start On Time" et "Start Off Time" avec couleur de texte spécifiée
    start_on_time_label = ctk.CTkLabel(home_frame, text="Start On Time", font=("Roboto", 18))
    start_on_time_label.pack(pady=20)
    
    def on_time_combobox_changed(heure):
        global selected_on_hour_led3, selected_off_hour_led3
        selected_on_hour_led3 = int(heure[:-1])
        selected_off_hour_led3 = (selected_on_hour_led3 + slider_value_led3) % 24
        off_time_combobox.set(f"{selected_off_hour_led3:02d}h")
    
    def off_time_combobox_changed(heure):
        global selected_on_hour_led3, selected_off_hour_led3
        selected_off_hour_led3 = int(heure[:-1])
        selected_on_hour_led3 = (selected_off_hour_led3 + slider_value2_led3) % 24
        on_time_combobox.set(f"{selected_on_hour_led3:02d}h")

    heures = [f"{i:02d}h" for i in range(24)]

    # Création des comboboxes et liaison des fonctions de callback
    on_time_combobox = ctk.CTkComboBox(home_frame, values=heures,command=on_time_combobox_changed)
    on_time_combobox.pack(pady=10)

    start_off_time_label = ctk.CTkLabel(home_frame, text="Start Off Time", font=("Roboto", 18))
    start_off_time_label.pack(pady=20)

    off_time_combobox = ctk.CTkComboBox(home_frame, values=heures,command=off_time_combobox_changed)
    off_time_combobox.pack(pady=10)

    slider.set(slider_value_led3)
    slider2.set(slider_value2_led3)
    on_time_combobox.set(f"{selected_on_hour_led3:02d}h")
    off_time_combobox.set(f"{selected_off_hour_led3:02d}h")
    slider_value_label.configure(text=f"{slider_value_led3} Hours")
    slider_value_label2.configure(text=f"{slider_value2_led3} Hours")

    # Création du bouton Save
    save_button = ctk.CTkButton(home_frame, text="Save",font=("Roboto", 24), command=save_values_led3)
    save_button.pack(anchor='e',padx=100)

     # Afficher l'heure actuelle
    global time_label
    time_label = ctk.CTkLabel(home_frame, font=("Roboto", 18))
    time_label.place(x=820)
    update_time()

def show_home():
    # Effacer tout contenu précédent dans home_frame
    for widget in home_frame.winfo_children():
        widget.destroy()

    # Ajouter le titre de la page d'accueil
    home_title = ctk.CTkLabel(home_frame, text="Aquaponic System Management", font=("Roboto", 36))
    home_title.pack(side="top", fill="x", pady=20)

    # Ajouter le texte descriptif
    description_text = ("This graphical users interface (GUI) is meticulously crafted to streamline the management "
                        "and control of a sophisticated, modular aquaponics system, which functions independently. "
                        "The interface boasts a user-friendly layout, allowing users to effortlessly adjust and "
                        "monitor the various automation parameters critical for the system's optimal performance. "
                        "Key features include the ability to regulate the on/off timing of the LED panels, which are "
                        "essential for the growth and nourishment of the plants. Additionally, users can precisely "
                        "control the flow rate of the water pump, ensuring the right amount of water circulation for "
                        "both plants and fish. The interface also facilitates the fine-tuning of dispensers for fish "
                        "food and chelated iron, crucial for maintaining the health of the fish and the overall "
                        "nutrient balance within the ecosystem. This comprehensive control system is designed to "
                        "ensure the smooth operation of the aquaponics system, promoting a sustainable and efficient "
                        "approach to urban farming.")

    description_label = ctk.CTkLabel(home_frame, text=description_text, font=("Roboto", 16), wraplength=500, justify="left")
    description_label.pack(anchor='w', padx=20, pady=60)

    # Chargement et redimensionnement de l'image du logo
    image = Image.open("Aquapo.png")  # Remplacez avec le chemin de votre image
    image = image.resize((400, 400))
    image = ImageTk.PhotoImage(image)

    # Création du label pour le logo
    image_label = ctk.CTkLabel(home_frame, image=image)
    image_label.image = image  # Garder une référence
    image_label.place(x=550,y=140)

     # Afficher l'heure actuelle
    global time_label
    time_label = ctk.CTkLabel(home_frame, font=("Roboto", 18))
    time_label.place(x=820)
    update_time()

def save_values_pump2():
    global saved_slider_value_pump2_on, saved_slider_value_pump2_off, slider_pump2, slider2_pump2
    saved_slider_value_pump2_on = slider_pump2.get()
    saved_slider_value_pump2_off = slider2_pump2.get()
    send_system4(saved_slider_value_pump2_on,saved_slider_value_pump2_off)

def show_pump2_management():
    global slider_pump2, slider2_pump2
    # Effacer tout contenu précédent dans home_frame
    for widget in home_frame.winfo_children():
        widget.destroy()

    # Ajouter le titre pour Pump 2 Management
    pump2_title = ctk.CTkLabel(home_frame, text="Pump 2 Management", font=("Roboto", 36))
    pump2_title.pack(pady=20)

    # Ajouter le texte pour le temps de la pompe
    on_time_label_pump2 = ctk.CTkLabel(home_frame, text="Pump Time ON", font=("Roboto", 18))
    on_time_label_pump2.pack(anchor='w', padx=180, pady=20)

   # Fonction pour mettre à jour le label en fonction de la valeur du slider
    def update_slider_pump2_label_on(value):
        minutes = int(value) // 2  # Convertir en minutes
        seconds = (int(value) % 2) * 30  # Rester les secondes
        slider_pump2_label.configure(text=f"{minutes} min {seconds:02d} sec")

    # Créer le label pour afficher la valeur du slider
    slider_pump2_label = ctk.CTkLabel(home_frame, text="0 min 00 sec", font=("Roboto", 18))
    slider_pump2_label.pack()

    # Créer le slider pour la pompe 2
    slider_pump2 = ctk.CTkSlider(home_frame, from_=0, to=20, number_of_steps=20, command=update_slider_pump2_label_on)
    slider_pump2.pack()

    # Ajouter le texte pour le temps de la pompe
    off_time_label_pump2 = ctk.CTkLabel(home_frame, text="Pump Time OFF", font=("Roboto", 18))
    off_time_label_pump2.pack(anchor='w', padx=180, pady=20)

   # Fonction pour mettre à jour le label en fonction de la valeur du slider
    def update_slider_pump2_label_off(value):
        minutes = int(value) // 2  # Convertir en minutes
        seconds = (int(value) % 2) * 30  # Rester les secondes
        slider2_pump2_label.configure(text=f"{minutes} min {seconds:02d} sec")

    # Créer le label pour afficher la valeur du slider
    slider2_pump2_label = ctk.CTkLabel(home_frame, text="0 min 00 sec", font=("Roboto", 18))
    slider2_pump2_label.pack()

    # Créer le slider pour la pompe 2
    slider2_pump2 = ctk.CTkSlider(home_frame, from_=0, to=20, number_of_steps=20, command=update_slider_pump2_label_off)
    slider2_pump2.pack()

    # Bouton Save pour Pump 2 Management
    save_button_pump2 = ctk.CTkButton(home_frame, text="Save", command=save_values_pump2, font=("Roboto", 24))
    save_button_pump2.pack(anchor='e', padx=100, pady=20)

    # Restaurer les valeurs sauvegardées
    slider_pump2.set(saved_slider_value_pump2_on)
    slider2_pump2.set(saved_slider_value_pump2_off)
    update_slider_pump2_label_on(saved_slider_value_pump2_on)
    update_slider_pump2_label_off(saved_slider_value_pump2_off)

     # Afficher l'heure actuelle
    global time_label
    time_label = ctk.CTkLabel(home_frame, font=("Roboto", 18))
    time_label.place(x=820)
    update_time()


# Initialisation de l'application
app = ctk.CTk()
app.geometry("1100x580")

# Définir le thème de couleur
ctk.set_default_color_theme("MoonlitSky.json")

# Création de la barre latérale
sidebar = ctk.CTkFrame(app, width=200, height=580)
sidebar.pack(side="left", fill="y")
sidebar.pack_propagate(False)  # Empêche le redimensionnement automatique

# Chargement et redimensionnement de l'image du logo
logo_image = Image.open("Logo.png")  # Remplacez avec le chemin de votre image
logo_image = logo_image.resize((100, 100))
logo_image = ImageTk.PhotoImage(logo_image)

# Création du label pour le logo
logo_label = ctk.CTkLabel(sidebar, image=logo_image)
logo_label.image = logo_image  # Garder une référence
logo_label.pack(pady=10)

# Création des boutons dans la barre latérale
button_a = ctk.CTkButton(sidebar, text="LED n°1", command=show_led1_management, width=180, height=40)
button_a.pack(pady=10)
# Ajoutez un bouton pour accéder à la page LED n°2 dans votre sidebar
button_b = ctk.CTkButton(sidebar, text="LED n°2", command=show_led2_management, width=180, height=40)
button_b.pack(pady=10)
# Ajoutez un bouton pour accéder à la page LED n°2 dans votre sidebar
button_c = ctk.CTkButton(sidebar, text="LED n°3", command=show_led3_management, width=180, height=40)
button_c.pack(pady=10)
# Ajoutez un bouton pour accéder à la page LED n°2 dans votre sidebar
button_d = ctk.CTkButton(sidebar, text="Pump", command=show_pump2_management, width=180, height=40)
button_d.pack(pady=10)

buttons = ["Distributor (Iron & Food)", "System Information","Parameters"]
for text in buttons:
    button = ctk.CTkButton(sidebar, text=text, width=180, height=40)
    button.pack(pady=10)

# Création du séparateur
separator_color = "#244461"
separator = ctk.CTkFrame(app, width=2, height=580, fg_color=separator_color)
separator.pack(side="left", fill="y")

# Création du cadre pour l'écran d'accueil
home_frame = ctk.CTkFrame(app)
home_frame.pack(side="right", fill="both", expand=True)

# Création du bouton pour le logo "Home"
home_button = ctk.CTkButton(sidebar, text="Home", command=show_home, width=90, height=40)
home_button.pack(side="bottom", pady=10)

# Afficher la page d'accueil au lancement du programme
show_home()

app.mainloop()